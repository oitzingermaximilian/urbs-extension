import os
import pandas as pd
import matplotlib.pyplot as plt
import glob
import openpyxl
import time
import seaborn as sns


def plot_nzia_benchmark(output_file_path):
    # Extract scenario name from file name
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")

    # Create output folder next to the Excel file
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    component_sheets = [
        "capacity_ext_stockout",
        "capacity_ext_euprimary",
        "capacity_ext_eusecondary",
    ]

    # Show actual sheet names
    wb = openpyxl.load_workbook(output_file_path, read_only=True)
    print(f"📄 Sheets in {output_file_path}: {wb.sheetnames}")
    # Read data
    data_frames = {
        sheet: pd.read_excel(output_file_path, sheet_name=sheet)
        for sheet in component_sheets
    }
    all_techs = sorted(
        set().union(*(df["key_1"].unique() for df in data_frames.values()))
    )

    years = [2025, 2030, 2035, 2040, 2045, 2050]

    for tech in all_techs:
        data = {}
        for sheet, df in data_frames.items():
            tech_df = df[df["key_1"] == tech].set_index("year").sort_index()
            series = tech_df["value"] / 1000  # Convert to GW
            data[sheet] = series.reindex(years).fillna(0)

        abs_data = pd.DataFrame(data)
        if abs_data.sum().sum() == 0:
            print(f"⚠ No data for technology '{tech}', skipping plots.")
            continue

        rel_data = abs_data.div(abs_data.sum(axis=1), axis=0).fillna(0)

        # === Styling ===
        colors = ["#FDC5B5", "#F99B7D", "#F76C5E"]  # Soft peach to coral
        hatches = ["..", "//", "xx"] #hatches = ["..", "//", "xx"]
        labels = ["Remanufacturing", "Stock", "Manufacturing"]
        sheet_order = [
            "capacity_ext_eusecondary",
            "capacity_ext_stockout",
            "capacity_ext_euprimary",
        ]

        # RELATIVE PLOT
        fig_rel, ax_rel = plt.subplots(figsize=(10, 6))
        bar_container = rel_data[sheet_order].plot(
            kind="bar",
            stacked=True,
            color=colors,
            ax=ax_rel,
            width=0.5,
            edgecolor="black",
            linewidth=0.5,
        )

        for bar_group, hatch in zip(bar_container.containers, hatches):
            for bar in bar_group:
                bar.set_hatch(hatch)

        ax_rel.set_title(
            f"Composition of Capacity Additions (40% Benchmark) - {tech}", pad=15
        )
        ax_rel.set_xlabel("Year", labelpad=10)
        ax_rel.set_xticks(range(len(years)))
        ax_rel.set_xticklabels(years)
        ax_rel.set_yticks([0, 0.25, 0.5, 0.75, 1.0])
        ax_rel.set_yticklabels(["0%", "10%", "20%", "30%", "40%"])
        ax_rel.set_ylim(0, 1.2)  # Create headroom
        ax_rel.legend(labels, frameon=False, bbox_to_anchor=(1, 1))
        ax_rel.grid(axis="y", alpha=0.3)

        # ABSOLUTE PLOT
        fig_abs, ax_abs = plt.subplots(figsize=(10, 6))
        bar_container_abs = abs_data[sheet_order].plot(
            kind="bar",
            stacked=True,
            color=colors,
            ax=ax_abs,
            width=0.5,
            edgecolor="black",
            linewidth=0.5,
        )

        for bar_group, hatch in zip(bar_container_abs.containers, hatches):
            for bar in bar_group:
                bar.set_hatch(hatch)

        ax_abs.set_title(f"Absolute Capacity Additions 40% Benchmark - {tech}", pad=15)
        ax_abs.set_ylabel("Capacity (GW)", labelpad=10)
        ax_abs.set_xlabel("Year", labelpad=10)
        ax_abs.set_xticks(range(len(years)))
        ax_abs.set_xticklabels(years)
        ax_abs.legend(labels, frameon=True, bbox_to_anchor=(1, 1))
        ax_abs.grid(axis="y", alpha=0.3)

        # Save figures
        safe_tech_name = tech.replace(" ", "_").replace("/", "_")
        fig_rel.savefig(
            os.path.join(output_dir, f"relative_composition_{safe_tech_name}.png"),
            dpi=300,
            bbox_inches="tight",
        )
        fig_abs.savefig(
            os.path.join(output_dir, f"absolute_capacity_{safe_tech_name}.png"),
            dpi=300,
            bbox_inches="tight",
        )

        plt.close(fig_rel)
        plt.close(fig_abs)

        print(f"✔ Plots saved for: {tech} in {output_dir}")

def plot_scrap(output_file_path):
    # Extract scenario name from file name
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")

    # Create output folder next to the Excel file
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    # Load scrap data
    sheet_name = "Total_Scrap"
    df = pd.read_excel(output_file_path, sheet_name=sheet_name)

    # Convert from tonnes to megatonnes (Mt)
    df["value"] = df["value"] / 1e6

    # Define the year range
    years = list(range(2024, 2051))

    # Get unique technologies
    techs = df["key_1"].unique()

    for tech in techs:
        tech_df = df[df["key_1"] == tech].set_index("year").sort_index()
        series = tech_df["value"].reindex(years).fillna(0)

        if series.sum() == 0:
            print(f"⚠ No data for technology '{tech}', skipping.")
            continue

        # Plot
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(series.index, series.values, color="seagreen", linewidth=2)  # smooth line

        ax.set_title(f"Scrap volume – {tech}")
        ax.set_xlabel("Year")
        ax.set_ylabel("Scrap [Mt]")
        ax.set_xlim(2023, 2051)
        #ax.set_xticks(range(2024, 2051, 2))  # Only show actual data years
        ax.set_ylim(0, max(series.values) * 1.1 if series.values.any() else 1)
        ax.grid(True, linestyle="--", alpha=0.3)

        plt.tight_layout()
        plot_filename = f"scrap_{tech}.png"
        fig.savefig(os.path.join(output_dir, plot_filename), dpi=300)
        plt.close(fig)

        print(f"✔ Plot saved for: {tech} → {plot_filename}")

def plot_installed_capacity(output_file_path):
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    # Load data
    df = pd.read_excel(output_file_path, sheet_name="Installed_Capacity_Q_s")

    # Filter out Batteries
    df = df[df["key_1"] != "Batteries"]

    # Pivot data
    pivot_df = df.pivot_table(index="year", columns="key_1", values="value", aggfunc="sum")
    pivot_df = pivot_df.sort_index()
    pivot_df = pivot_df / 1000  # MW to GW
    years = [2025, 2030, 2035, 2040, 2045, 2050]
    pivot_df = pivot_df.reindex(years).fillna(0)

    # Sort columns consistently
    techs = sorted(pivot_df.columns)
    pivot_df = pivot_df[techs]

    # Generate a beautiful qualitative color palette
    n_techs = len(techs)
    colors = sns.color_palette("Set3", n_colors=n_techs)


    # === Absolute plot ===
    fig_abs, ax_abs = plt.subplots(figsize=(11, 6))
    pivot_df.plot(
        kind="bar",
        stacked=True,
        color=colors,
        ax=ax_abs,
        edgecolor="black",
        linewidth=0.3,
        width=0.6
    )

    ax_abs.set_title("Absolute Installed Capacity (excl. Batteries)", pad=15)
    ax_abs.set_ylabel("Capacity (GW)")
    ax_abs.set_xlabel("Year")
    ax_abs.set_xticks(range(len(years)))
    ax_abs.set_xticklabels(years)
    ax_abs.grid(axis="y", alpha=0.3)
    ax_abs.legend(title="Technology", bbox_to_anchor=(1.02, 1), loc="upper left")

    fig_abs.tight_layout()
    fig_abs.savefig(
        os.path.join(output_dir, "installed_capacity_absolute.png"),
        dpi=300,
        bbox_inches="tight"
    )

    # === Relative plot ===
    rel_df = pivot_df.div(pivot_df.sum(axis=1), axis=0).fillna(0)

    fig_rel, ax_rel = plt.subplots(figsize=(11, 6))
    rel_df.plot(
        kind="bar",
        stacked=True,
        color=colors,
        ax=ax_rel,
        edgecolor="black",
        linewidth=0.3,
        width=0.6
    )

    ax_rel.set_title("Relative Installed Capacity Share (excl. Batteries)", pad=15)
    ax_rel.set_ylabel("Share of Total Capacity")
    ax_rel.set_xlabel("Year")
    ax_rel.set_xticks(range(len(years)))
    ax_rel.set_xticklabels(years)
    ax_rel.set_yticklabels(["{:.0%}".format(y) for y in ax_rel.get_yticks()])
    ax_rel.grid(axis="y", alpha=0.3)
    ax_rel.legend(title="Technology", bbox_to_anchor=(1.02, 1), loc="upper left")

    fig_rel.tight_layout()
    fig_rel.savefig(
        os.path.join(output_dir, "installed_capacity_relative.png"),
        dpi=300,
        bbox_inches="tight"
    )

    plt.close(fig_abs)
    plt.close(fig_rel)

    print(f"✔ Installed capacity plots saved in {output_dir}")


# Example: Call this after saving Excel
# write_carryovers_to_excel(..., output_file_path)
# plot_capacity_decomposition_by_technology(output_file_path)
def plot_all_scenarios(base_dir): #TODO re-add if needed
    # Find all Excel files in the base_dir and its subfolders
    excel_files = glob.glob(
        os.path.join(base_dir, "**", "result_scenario_*.xlsx"), recursive=True
    )

    for file in excel_files:
        print(f"📊 Processing: {file}")
        expected_sheets = [
            "Installed_Capacity_Q_s",
            "Existing_Stock_Q_stock",
            "capacity_dec_start",
            "Total Cap Sec",
            "CO2_emissions",
            "Total_Cost",
            "Total_Scrap",
            "Pricereduction",
            "Balance",
            "Commodities_Demand",
            "capacity_ext_imported",
            "capacity_ext_stockout",
            "capacity_ext_euprimary",
            "capacity_ext_eusecondary",
            "capacity_ext_stock_imported",
            "newly_added_capacity",
        ]

        wait_for_excel_sheets(file, expected_sheets)
        plot_capacity_decomposition_by_technology(file)


def wait_for_excel_sheets(path, expected_sheets, timeout=60): #TODO re-add if needed
    """Wait until the expected sheets exist in the Excel file."""
    start = time.time()
    while time.time() - start < timeout:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            if all(sheet in sheets for sheet in expected_sheets):
                return True
        except Exception:
            pass
        time.sleep(0.5)
    raise TimeoutError(f"Expected sheets not found in {path} after {timeout} seconds.")


# plot_nzia_benchmark("result/urbs-20250520T1651/result_scenario_base.xlsx")
# plot_all_scenarios("result")
# plot_scrap("result/urbs-20250520T1651/result_scenario_base.xlsx")
# plot_installed_capacity("result/urbs-20250520T1651/result_scenario_base.xlsx")